"""从 COC7 规则表提取职业原始数据（issue #114 第一步：确定性提取）。

这一步**只做搬运，不做任何解释**——把两个 sheet 里跟职业有关的原始内容原样
抄进 JSON，解释工作留给下一步。这样划分是因为两件事的可靠性完全不同：搬运
是确定性的、可重复的；解释需要判断（见下面"为什么不能直接机读"）。

产物 `raw_occupations.json` **进 git**。规则表本身不在仓库里
（`/Users/apple/Developer/work/rules/COC7空白卡CY23Final.xlsx`，因体积和版权
原因不入库），所以如果只提交最终的 `coc7_content.py`，任何人都无法在仓库内
复核那 229 条数据是怎么来的。提交原始 JSON 才让推导过程可审计。

## 为什么不能直接机读矩阵

`本职技能` sheet 是 230 列职业 × 66 行技能的矩阵，乍看结构清晰：

- 符号行 `☆ ⊙ ☯ ※ 任意特长` 编码自选槽数量，图例在 `更新说明` sheet：
  `☆ 二选一`、`⊙ 二选一`、`☯ 社交技能`、`※ 多选，或X选一`
- 技能行里 `★` 标记固定本职技能

但**专精格是给人看的显示残片，不是数据**：

- 艺人 `技艺①='表' ②='演' ③='类'` —— "表演类"三个字被摊在三个格子里
- 秘书 `技艺②='字☆' ③='记☆'` —— 原文是「技艺（打字或速记）」，首字被截掉
- 绅士、淑女 `射击①='步/霰'` —— 一格多值，**且不带 `★`**（只认 `★` 会把整条丢掉）

而且两个 sheet 会互相矛盾：艺人的自然语言里有「闪避」「母语」，矩阵里都没有。
225 个可比对的职业中有 106 个（47%）含这类残片。

所以这里两个 sheet 都原样导出，由下一步交叉校验、逐条裁定。

## 运行

`openpyxl` 不是本项目的依赖（CI 里根本没有规则表可读，装了也是白装），
用任何装了 openpyxl 的解释器跑即可：

    python3 scripts/coc7_occupations/extract_raw.py [规则表路径]
"""

import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSX = Path("/Users/apple/Developer/work/rules/COC7空白卡CY23Final.xlsx")
OUTPUT = Path(__file__).parent / "raw_occupations.json"

# `职业列表` sheet 的列位置（0-based）。表头在第 0 行，数据从第 1 行起。
COL_INDEX, COL_NAME, COL_CREDIT, COL_FORMULA, COL_SKILLS = 0, 1, 3, 4, 6
# 职业介绍（`OccupationSpec.description` 的来源）和推荐关系人。后者本期不入库，
# 但一并导出——它是建卡时给玩家写背景的素材，属于同一条数据的一部分，将来要用
# 时不必再回头改提取器重跑。
COL_CONTACTS, COL_DESCRIPTION = 10, 12

# `本职技能` sheet 里的自选槽符号行（行号 → 符号），见模块文档的图例。
SLOT_ROWS = {2: "☆", 3: "⊙", 4: "☯", 5: "※", 6: "任意特长"}

# 序号 0 是使用说明行、序号 1 是"自定义职业"（不是真实职业，本期不导入），
# 真实职业从序号 2 起。
FIRST_REAL_INDEX = 2

# 两个 sheet 里写法不同、但确认是同一个职业的名称对照（`职业列表` 名 → `本职技能` 名）。
#
# 每一条都逐项比对过两边内容才认定，证据记在下面。**不做模糊匹配**——名字相近
# 的职业在这张表里大量存在（`科学家`/`司法家`/`科学搜查研究员` 三个都在），
# 按相似度自动挑一个迟早会挑错，而且错了没有任何信号。
NAME_ALIASES = {
    # 原文「一项社交技能，格斗（斗殴），射击，法律，图书馆，心理学，潜行，追踪」
    # 与矩阵 `事务所侦探`（☯=1；格斗/射击/法律/图书馆使用/心理学/潜行/追踪 均 ★）逐项吻合。
    "事务所侦探、保安": "事务所侦探",
    # 原文「汽车驾驶，电气维修，格斗，急救，机械维修，操作重型机械，投掷，任意一项」
    # 与矩阵 `非熟练工人`（任意特长=1；其余 7 项均 ★）逐项吻合。
    "工人-非熟练工人": "非熟练工人",
    # 原文「艺术（摄影），医学，法律，科学（化学，司法科学，药学），侦察，任意一项」
    # 与矩阵 `司法家`（技艺①=摄影；科学①=化学 ②=司法 ③=药学；法律/医学/侦查 ★；
    # 任意特长=1）逐项吻合——`司法家` 是 `司法科学家` 的截断显示。
    "司法科学家": "司法家",
}


def _match_key(name: str) -> str:
    """按名字匹配时用的键：去掉所有空白。

    只在序号对不上时兜底用（见 `extract`）。只规范化空白——这属于搬运不属于
    解释（`赛车手/ 赛艇手` 与 `赛车手/赛艇手` 显然是同一个职业）。**不做模糊
    匹配**：名称真正不同的条目要留给下一步裁定，静默挑一个"最像的"会把需要人
    看的信号藏起来。
    """
    return "".join(name.split())


def extract(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    prose_rows = list(wb["职业列表"].iter_rows(values_only=True))
    matrix_rows = list(wb["本职技能"].iter_rows(values_only=True))
    skill_labels = [r[0] for r in matrix_rows]
    matrix_names = list(matrix_rows[1])

    # 🔴 按**序号**对齐两个 sheet，不按名字。
    #
    # `职业列表` 里有 6 对同名职业（艺术家/工匠/艺人/律师/私家侦探/科学家 各两条，
    # 是同名不同版本的设定），矩阵里也相应有两列（如 `私家侦探` 在第 92 和 220 列）。
    # 第一版用 {名字: 列号} 的字典匹配，同名后写覆盖先写，每一对里都有一条拿到了
    # 另一条的矩阵数据——而且**悄无声息**。
    #
    # 矩阵第 0 行就是序号，两个 sheet 的序号是同一套编号，按它对齐是精确的。
    # 名字匹配只在序号对不上时兜底（`NAME_ALIASES` 那三条属于这种）。
    matrix_by_index = {
        matrix_rows[0][i]: i for i in range(len(matrix_names)) if isinstance(matrix_rows[0][i], int)
    }
    matrix_by_name = {_match_key(n): i for i, n in enumerate(matrix_names) if isinstance(n, str)}

    occupations = []
    for row in prose_rows[1:]:
        index, name = row[COL_INDEX], row[COL_NAME]
        if not isinstance(index, int) or index < FIRST_REAL_INDEX or not name:
            continue

        entry = {
            "index": index,
            "name": name,
            "credit_raw": row[COL_CREDIT],
            "formula_raw": row[COL_FORMULA],
            "prose_skills": row[COL_SKILLS],
            "description": row[COL_DESCRIPTION],
            "contacts": row[COL_CONTACTS],
            # 矩阵里对不上号时留空，而不是猜一个最相近的列——名称对不上本身
            # 就是需要人看的信号，静默匹配会把它藏起来。
            "matrix_column": None,
            "matrix_slots": {},
            "matrix_skills": {},
        }

        col = matrix_by_index.get(index)
        if col is None:
            col = matrix_by_name.get(_match_key(NAME_ALIASES.get(name, name)))
        if col is not None:
            entry["matrix_column"] = col
            for row_no, symbol in SLOT_ROWS.items():
                value = matrix_rows[row_no][col] if col < len(matrix_rows[row_no]) else None
                if value not in (None, 0, ""):
                    entry["matrix_slots"][symbol] = value
            for i, matrix_row in enumerate(matrix_rows):
                if i <= max(SLOT_ROWS) or skill_labels[i] in (None, 0):
                    continue
                value = matrix_row[col] if col < len(matrix_row) else None
                if value not in (None, 0, ""):
                    entry["matrix_skills"][str(skill_labels[i])] = str(value)

        occupations.append(entry)

    return occupations


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f"规则表不存在: {xlsx_path}")

    occupations = extract(xlsx_path)
    OUTPUT.write_text(
        json.dumps(occupations, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    unmatched = [o["name"] for o in occupations if o["matrix_column"] is None]
    print(f"提取 {len(occupations)} 个职业 -> {OUTPUT}")
    print(f"在 `本职技能` 矩阵里找不到对应列的: {len(unmatched)}")
    for name in unmatched:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
